"""
╔══════════════════════════════════════════════════════════════╗
║   EXCEL WRITER  —  Xuất file 20 cột có công thức & style     ║
║   Format theo file mẫu AGlobal Keyword Research              ║
╚══════════════════════════════════════════════════════════════╝
"""
import os
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.constants import (
    OUTPUT_COLS, FULL_EXPORT_COLS, FULL_EXPORT_WIDTHS, COLORS_ARGB, OUTPUT_DIR
)


class ExcelWriter:
    """Xuất DataFrame đã lọc ra file Excel 20 cột có công thức."""

    @staticmethod
    def format_output_name(prefix: str = "PPC_Keywords") -> str:
        """Sinh tên file theo format: {prefix}_{HH:MM:SS}_{DD-MM-YYYY}.xlsx"""
        now = datetime.now()
        time_part = now.strftime("%H:%M:%S")
        date_part = now.strftime("%d-%m-%Y")
        return f"{prefix}_{time_part}_{date_part}.xlsx"

    # ── Export 6 cột (giữ lại cho preview UI) ────────────────────────────
    def export_6cols(
        self,
        df: pd.DataFrame,
        output_path: str,
        brands: list[str] | None = None,
    ) -> tuple[str, list]:
        """Xuất 6 cột chuẩn (giữ lại tương thích cũ)."""
        return self.export_full(df, output_path, brands, price=None, amazon_fee=None)

    # ── Export 20 cột với công thức Excel ────────────────────────────────
    def export_full(
        self,
        df: pd.DataFrame,
        output_path: str,
        brands: list[str] | None = None,
        price: float | None = None,
        amazon_fee: float | None = None,
    ) -> tuple[str, list]:
        """
        Xuất 20 cột chuẩn (A-T) ra file Excel với công thức.
        Format theo file mẫu AGlobal Keyword Research.

        Các cột công thức:
          H (Clicks)        = G{n} * B{n}        (CTR * Search Volume)
          I (Spend)         = F{n} * H{n}        (Bid * Clicks)
          K (Ads Orders)    = H{n} * J{n}        (Clicks * CVR)
          M (Ads Revenue)   = K{n} * L{n}        (Ads Orders * Price)
          N (ACOS)          = I{n} / M{n}        (Spend / Ads Revenue)
          O (Total Orders)  = K{n} + K{n}/4      (Ads + organic)
          P (Total Revenue) = L{n} * O{n}        (Price * Total Orders)
          Q (Product Fee)   = L{n} / 3           (Price / 3)
          S (Total Fee)     = (R{n}+Q{n})*O{n}+I{n}
          T (Profit)        = P{n} - S{n}

        Trả về: (đường dẫn file, danh sách brand keywords đã highlight)
        """
        if brands is None:
            brands = []

        n_rows = len(df)

        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        # ── Kiểu dáng (không border, không xen kẽ — theo file mẫu) ───
        header_fill = PatternFill("solid", start_color=COLORS_ARGB["header_bg"])
        header_font = Font(name='Arial', bold=True, color=COLORS_ARGB["header_fg"], size=10)
        header_align = Alignment(horizontal='center', vertical='bottom')

        data_font = Font(name='Arial', size=10)
        data_font_bold = Font(name='Arial', size=10, bold=True)
        data_font_red = Font(name='Arial', size=10, color="FFFF0000")       # CTR & CVR
        data_font_bold_red = Font(name='Arial', size=10, bold=True, color="FFFF0000")
        brand_fill = PatternFill("solid", start_color=COLORS_ARGB["brand_warn"])

        # ── Headers (Row 1) ────────────────────────────────────────────
        headers = [col[0] for col in FULL_EXPORT_COLS]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(1, ci, h)
            c.fill = header_fill
            c.font = header_font
            c.alignment = header_align
            width = FULL_EXPORT_WIDTHS.get(h, 14)
            ws.column_dimensions[get_column_letter(ci)].width = width
        ws.row_dimensions[1].height = 22

        # ── Data rows ──────────────────────────────────────────────────
        brands_found = []
        kw_col = "Keyword Phrase"

        for ri in range(n_rows):
            rn = ri + 2  # Excel row number

            # Lấy keyword để check brand
            keyword = str(df.iloc[ri][kw_col]).lower() if kw_col in df.columns else ""
            is_brand = any(b.lower() in keyword for b in brands) if brands else False
            if is_brand:
                brands_found.append(df.iloc[ri][kw_col])

            for ci, (header, src_col, default, num_fmt) in enumerate(FULL_EXPORT_COLS, 1):
                c = ws.cell(rn, ci)

                # ── Xác định giá trị ──────────────────────────────────
                if src_col and src_col in df.columns:
                    # Lấy từ dữ liệu Cerebro
                    raw = df.iloc[ri][src_col]
                    if pd.isna(raw):
                        val = None
                    elif header == "Bid":
                        val = round(float(raw), 2)
                    elif header in ("Search Volume", "Sponsored ASINs", "Competing Products", "CPR"):
                        try:
                            val = int(float(raw))
                        except (ValueError, TypeError):
                            val = raw
                    else:
                        val = raw
                elif header == "Price" and price is not None:
                    val = round(price, 2)
                elif header == "Amazon Fee" and amazon_fee is not None:
                    val = round(amazon_fee, 2)
                elif default is not None:
                    if isinstance(default, str) and "{" in default:
                        # Công thức Excel: thay {n} bằng số dòng thực
                        val = default.replace("{n}", str(rn))
                    else:
                        val = default
                else:
                    val = None

                # ── Ghi vào cell ──────────────────────────────────────
                if val is not None:
                    c.value = val
                # Nếu val là None, để trống (Price, Amazon Fee)

                # ── Style ─────────────────────────────────────────────
                if header == "Keyword Phrase":
                    c.font = data_font_bold
                elif header in ("CTR", "CVR"):
                    c.font = data_font_red
                else:
                    c.font = data_font
                c.alignment = Alignment(vertical='bottom')

                # Brand highlight (vàng) — theo đúng file mẫu, không xen kẽ màu
                if is_brand and brands:
                    c.fill = brand_fill

                # Number format
                if num_fmt:
                    c.number_format = num_fmt

            ws.row_dimensions[rn].height = 18

        # ── Total row ──────────────────────────────────────────────────
        total_row = n_rows + 2
        last_data_row = total_row - 1
        total_fill = PatternFill("solid", start_color="FFFFE599")
        total_font = Font(name='Arial', size=10, bold=True)

        # Column mapping cho dòng Total
        total_defs = {
            1:  ("Total", None),
            6:  (f"=AVERAGE(F2:F{last_data_row})", '"$"#,##0.00'),
            9:  (f"=SUM(I2:I{last_data_row})", '"$"#,##0.00'),
            11: (f"=SUM(K2:K{last_data_row})", '#,##0'),
            13: (f"=SUM(M2:M{last_data_row})", '"$"#,##0.00'),
            15: (f"=SUM(O2:O{last_data_row})", '#,##0'),
            16: (f"=SUM(P2:P{last_data_row})", '"$"#,##0.00'),
            19: (f"=SUM(S2:S{last_data_row})", '"$"#,##0.00'),
            20: (f"=SUM(T2:T{last_data_row})", '"$"#,##0.00'),
        }

        for ci in range(1, 21):
            c = ws.cell(total_row, ci)
            if ci in total_defs:
                val, num_fmt = total_defs[ci]
                c.value = val
                if num_fmt:
                    c.number_format = num_fmt
            c.fill = total_fill
            c.font = total_font
            c.alignment = Alignment(vertical='bottom')

        ws.row_dimensions[total_row].height = 20

        # ── Campaign section rows (orange, giữa Total và Total thêm) ────
        section_fill = PatternFill("solid", start_color=COLORS_ARGB["header_bg"])
        section_font = Font(name='Arial', size=10, bold=True, color=COLORS_ARGB["header_fg"])
        section_align = Alignment(horizontal='left', vertical='bottom')

        section_rows = [
            ("Campaigns từ khóa ngắn",       0.05),
            ("Auto",                         0.20),
            ("Cate",                         0.10),
            ("ASIN",                         0.05),
            ("Từ khóa dạng phrase + broad",  0.20),
        ]

        sect_start_row = total_row + 1
        for si, (sect_name, ctr_val) in enumerate(section_rows):
            sr = sect_start_row + si
            for ci in range(1, 21):
                c = ws.cell(sr, ci)
                c.fill = section_fill
                c.font = section_font
                c.alignment = section_align
            # Column A: section name
            ws.cell(sr, 1).value = sect_name
            # Column B: CTR value (đỏ)
            ctr_cell = ws.cell(sr, 2)
            ctr_cell.value = ctr_val
            ctr_cell.number_format = "0.0%"
            ctr_cell.font = data_font_bold_red
            ws.row_dimensions[sr].height = 20

        sect_end_row = sect_start_row + len(section_rows) - 1

        # ── Total thêm 50-100% (×160%) ────────────────────────────────
        total2_row = sect_end_row + 1
        total2_defs = {
            1:  ("Total thêm 50 - 100%", None),
            2:  (f"=SUM(B{sect_start_row}:B{sect_end_row})", "0.0%"),
            9:  (f"=I{total_row}*160%", '"$"#,##0.00'),
            11: (f"=K{total_row}*160%", '#,##0'),
            13: (f"=M{total_row}*160%", '"$"#,##0.00'),
            15: (f"=O{total_row}*160%", '#,##0'),
            16: (f"=P{total_row}*160%", '"$"#,##0.00'),
            19: (f"=S{total_row}*160%", '"$"#,##0.00'),
            20: (f"=T{total_row}*160%", '"$"#,##0.00'),
        }

        for ci in range(1, 21):
            c = ws.cell(total2_row, ci)
            if ci in total2_defs:
                val, num_fmt = total2_defs[ci]
                c.value = val
                if num_fmt:
                    c.number_format = num_fmt
            c.fill = total_fill
            c.font = data_font_bold_red if ci == 2 else total_font
            c.alignment = Alignment(vertical='bottom')

        ws.row_dimensions[total2_row].height = 20

        # ── Freeze pane ────────────────────────────────────────────────
        ws.freeze_panes = "A2"

        # ── Save ───────────────────────────────────────────────────────
        save_dir = os.path.dirname(output_path)
        if save_dir:
            os.makedirs(save_dir, exist_ok=True)
        wb.save(output_path)

        return output_path, brands_found
